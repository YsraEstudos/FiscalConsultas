"""Análise da estrutura do arquivo tipi.xlsx para entender colunas e dados."""
import openpyxl

wb = openpyxl.load_workbook('data/tipi.xlsx')
ws = wb.active

print("=== PRIMEIRAS 30 LINHAS DO tipi.xlsx ===")
print(f"{'Row':>5} | {'Col A (NCM)':^20} | {'Col B (Ex)':^8} | {'Col C (Descricao)':^50} | {'Col D (Aliq)':^10}")
print("-" * 110)

for row_num in range(1, 31):
    a = ws.cell(row=row_num, column=1).value or ''
    b = ws.cell(row=row_num, column=2).value or ''
    c = ws.cell(row=row_num, column=3).value or ''
    d = ws.cell(row=row_num, column=4).value or ''
    
    # Truncar descrição
    c_trunc = str(c)[:48] + '..' if len(str(c)) > 50 else str(c)
    print(f"{row_num:>5} | {str(a):^20} | {str(b):^8} | {c_trunc:<50} | {str(d):^10}")

print("\n=== LINHAS DO CAPÍTULO 84.13 ===")
found = 0
for row_num in range(1, ws.max_row + 1):
    a = str(ws.cell(row=row_num, column=1).value or '')
    if '8413' in a or a.startswith('84.13'):
        b = ws.cell(row=row_num, column=2).value or ''
        c = ws.cell(row=row_num, column=3).value or ''
        d = ws.cell(row=row_num, column=4).value or ''
        c_trunc = str(c)[:48] + '..' if len(str(c)) > 50 else str(c)
        print(f"{row_num:>5} | {a:^20} | {str(b):^8} | {c_trunc:<50} | {str(d):^10}")
        found += 1
        if found >= 40:
            break

wb.close()
